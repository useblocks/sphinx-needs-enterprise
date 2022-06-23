import logging
import operator
import os
from functools import reduce  # forward compatibility for Python 3
from typing import Dict, List

from jinja2 import Template
from openpyxl import load_workbook

log = logging.getLogger(__name__)


def dict_get(root, items, default=None):
    """
    Access a nested object in root by item sequence.

    Usage::
       data = {"nested": {"a_list": [{"finally": "target_data"}]}}
       value = dict_get(["nested", "a_list", 0, "finally"], "Not_found")

    """
    try:
        value = reduce(operator.getitem, items, root)
    except (KeyError, IndexError, TypeError) as e:
        log.debug(e)
        return default
    return value


def dict_undefined_set(dict_obj, key, value):
    if key not in dict_obj:
        dict_obj[key] = value


def get_excel_data(
    file_path: str, end_row: int, end_col: int, start_row: int = 2, start_col: int = 1, header_row: int = 1
):
    if not os.path.exists(file_path):
        raise ReferenceError(f"Could not load spreadsheet file {file_path}")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    data = list(ws.iter_rows(min_col=start_col, max_col=end_col, min_row=start_row, max_row=end_row, values_only=True))
    column_header = list(
        ws.iter_rows(min_col=start_col, max_col=end_col, min_row=header_row, max_row=header_row, values_only=True)
    )[0]
    finalised_data: List[Dict] = []
    for row in data:
        output = {str(key): value for key, value in zip(column_header, row)}
        finalised_data.append(output)

    # Close the workbook after reading
    wb.close()
    return finalised_data


def jinja_parse(context: Dict, jinja_string: str) -> str:
    """
    Function to parse mapping options set to a string containing jinja template format.

    :param context: Data to be used as context in rendering jinja template
    :type: dict
    :param jinja_string: A jinja template string
    :type: str
    :return: A rendered jinja template as string
    :rtype: str

    """
    try:
        content_template = Template(jinja_string, autoescape=True)
    except Exception as e:
        raise ReferenceError(f'There was an error in the jinja statement: "{jinja_string}". ' f"Error Msg: {e}")

    content = content_template.render(**context)
    return content


def filter_excel_data(context: List[Dict], filter_string: str) -> List[Dict]:
    """
    Function to filter Excel service data and return the filtered data.

    :param context: Data imported from the Excel service
    :type: List[Dict]
    :param filter_string: A string containing Python statements
    :type: str
    :return: Filtered data
    :rtype: List[Dict]

    """
    data = context
    query = filter_string

    filtered_data: List[Dict] = []
    for row in data:
        try:
            if eval(query, {}, row) is True:
                filtered_data.append(row)
        except Exception as e:
            raise ReferenceError(f'There was an error in the query statement: "{filter_string}". ' f"Error Msg: {e}")

    return filtered_data
